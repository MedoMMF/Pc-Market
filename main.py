from tkinter import *
from tkinter import ttk
import datetime
from tkinter import messagebox
import webbrowser
import openpyxl
from openpyxl import Workbook
import os
import sys

pro = Tk()
pro.geometry('800x450+280+50')
pro.resizable(False, False)
pro.title('PC_MARKET')
pro.iconbitmap('D:\pcrepair.ico')


def open():
    root = Toplevel(bg='gray')
    root.geometry('950x552+280+50')
    root.resizable(False, False)
    root.iconbitmap('D:\pcrepair.ico')
    root.title('Pc_Market   (متجر للحاسوب الشخصي)')
    now = datetime.datetime.now()

    date = now.strftime("%y-%m-%d")

    Wb = Workbook()
    ws = Wb.active
    ws.title = 'Customer'
    ws["A1"] = 'Name'
    ws["B1"] = 'Number phone'
    ws["C1"] = 'address'
    ws["D1"] = 'Total'
    ws["E1"] = 'Date Buy'
    Wb.save('PcMarket.xlsx')

    menu = {
        0: ['شاشة كمبيوتر', 15000],
        1: ['كيبورد', 350],
        2: ['ميكرفون', 500],
        3: ['ماوس', 200],
        4: ['سماعة رأس', 400],
        5: ['ماوس باد', 50],
        6: ['ماوس لاسلكي', 250],
        7: ['سمعات مكبرا', 500]
    }

    def bill():
        global En_name
        global En_phone
        global En_add
        global En_total
        global En_date

        root.geometry('1205x552')
        F4 = Frame(root, bg='#5F7161', width=250,
                   height=434, bd=2, relief=GROOVE)
        F4.place(x=950, y=1)
        L_name = Label(F4, text='اسم المشتري', bg='#5F7161', fg='white')
        L_name.place(x=168, y=10)
        En_name = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
        En_name.place(x=15, y=40)

        L_phone = Label(F4, text='رقم المشتري', bg='#5F7161', fg='white')
        L_phone.place(x=170, y=70)
        En_phone = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
        En_phone.place(x=15, y=100)

        L_add = Label(F4, text='عنوان المشتري', bg='#5F7161', fg='white')
        L_add.place(x=160, y=130)
        En_add = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
        En_add.place(x=15, y=160)

        L_total = Label(F4, text='الحساب الكلي', bg='#5F7161', fg='white')
        L_total.place(x=165, y=190)
        En_total = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
        En_total.place(x=15, y=210)

        L_date = Label(F4, text='تاريخ الشراء', bg='#5F7161', fg='white')
        L_date.place(x=175, y=240)
        En_date = Entry(F4, width=24, font=('Tajawal', 12), justify=CENTER)
        En_date.place(x=15, y=270)
        add_buttn = Button(F4, text='حفظ الفاتوره',
                           width=31, cursor='hand2', bg='#EDDBC0', command=Wb.active)
        add_buttn.place(x=12, y=310)

        add_buttn = Button(F4, text='افراغ الحقول',
                           width=31, cursor='hand2', bg='#EDDBC0', command=clear1)
        add_buttn.place(x=12, y=340)

        add_buttn = Button(F4, text='بحث عن مشتري',
                           width=31, cursor='hand2', bg='#EDDBC0')
        add_buttn.place(x=12, y=370)

        add_buttn = Button(F4, text='حذف فاتوره',
                           width=31, cursor='hand2', bg='#EDDBC0')
        add_buttn.place(x=12, y=400)
        total = 0
        for item in trv.get_children():
            trv.delete(item)
        for i in range(len(sb)):
            if (int(sb[i].get()) > 0):
                price = int(sb[i].get())*menu[i][1]
                total = total+price
                myst = (str(menu[i][1]), str(sb[i].get()), str(price))
                trv.insert('', 'end', iid=i, text=menu[i][0], values=myst)
        finall = total
        En_total.insert('1', str(finall) + '$')
        En_date.insert('1', str(date))
        En_name.insert('1', str(name1[0]))
        En_phone.insert('1', str(name1[1]))
        En_add.insert('1', str(name1[2]))

    def clear():
        for item in trv.get_children():
            trv.delete(item)

        En_name.delete('0', END)
        En_phone.delete('0', END)
        En_add.delete('0', END)
        En_total.delete('0', END)
        En_date.delete('0', END)

    def clear1():
        En_name.delete('0', END)
        En_phone.delete('0', END)
        En_add.delete('0', END)
        En_total.delete('0', END)
        En_date.delete('0', END)

    F1 = Frame(root, bg='silver', width=600, height=550)
    F1.place(x=1, y=1)

    img_menu1 = PhotoImage(file='marketphoto/1.png')
    img_menu2 = PhotoImage(file='marketphoto/2.png')
    img_menu3 = PhotoImage(file='marketphoto/3.png')
    img_menu4 = PhotoImage(file='marketphoto/4.png')
    img_menu5 = PhotoImage(file='marketphoto/5.png')
    img_menu6 = PhotoImage(file='marketphoto/6.png')
    img_menu7 = PhotoImage(file='marketphoto/7.png')
    img_menu8 = PhotoImage(file='marketphoto/8.png')

    title = Label(F1, text='متجر للحاسب الالي', font=(
        'tajawal', 13), fg='white', bg='#5F7161', width=70)
    title.place(x=0, y=0)

    menu1 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu1, text='شاشة كمبيوتر', compound=TOP)
    menu1.place(x=40, y=45)

    menu2 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu2, text='كيبورد', compound=TOP)
    menu2.place(x=170, y=45)

    menu3 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu3, text='ميكرفون', compound=TOP)
    menu3.place(x=300, y=45)

    menu4 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu4, text='ماوس', compound=TOP)
    menu4.place(x=430, y=45)

    menu5 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu5, text='سماعة رأس', compound=TOP)
    menu5.place(x=40, y=180)

    menu6 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu6, text='ماوس باد', compound=TOP)
    menu6.place(x=170, y=180)

    menu7 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu7, text='ماوس لاسلكي', compound=TOP)
    menu7.place(x=300, y=180)

    menu8 = Button(F1, width=88, bg='white', bd=1, relief=SOLID, cursor='hand2',
                   height=85, image=img_menu8, text='سماعات مكبرا', compound=TOP)
    menu8.place(x=430, y=180)

    sb = []
    font1 = ('Times', 12, 'normal')
    sv1 = IntVar()
    sv2 = IntVar()
    sv3 = IntVar()
    sv4 = IntVar()
    sv5 = IntVar()
    sv6 = IntVar()
    sv7 = IntVar()
    sv8 = IntVar()

    sb1 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv1)
    sb1.place(x=40, y=140)
    sb.append(sb1)

    sb2 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv2)
    sb2.place(x=170, y=140)
    sb.append(sb2)

    sb3 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv3)
    sb3.place(x=300, y=140)
    sb.append(sb3)

    sb4 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv4)
    sb4.place(x=430, y=140)
    sb.append(sb4)

    sb5 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv5)
    sb5.place(x=40, y=275)
    sb.append(sb5)

    sb6 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv6)
    sb6.place(x=170, y=275)
    sb.append(sb6)

    sb7 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv7)
    sb7.place(x=300, y=275)
    sb.append(sb7)

    sb8 = Spinbox(F1, from_=0, to_=5, font=font1,
                  width=10, textvariable=sv8)
    sb8.place(x=430, y=275)
    sb.append(sb8)

    b1 = Button(F1, text='شراء الموارد', fg='white',
                font=('Tajawal,12'), width=15, bg='#6D8B74', bd=1, relief=SOLID, cursor='hand2', height=1, command=bill)
    b1.place(x=40, y=500)

    b2 = Button(F1, text='فاتوره جديده', fg='white',
                font=('Tajawal,12'), width=15, bg='#6D8B74', bd=1, relief=SOLID, cursor='hand2', height=1, command=clear)
    b2.place(x=200, y=500)

    b3 = Button(F1, text='اغلاق البرنامج', fg='white',
                font=('Tajawal,12'), width=15, bg='red', bd=1, relief=SOLID, cursor='hand2', height=1, command=quit)
    b3.place(x=390, y=500)

    F2 = Frame(root, bg='gray', width=343, height=550)
    F2.place(x=604, y=1)

    trv = ttk.Treeview(F2, selectmode='browse')
    trv.place(x=1, y=1, width=340, height=550)

    trv['columns'] = ['1', '2', '3']
    trv.column("#0", width=80, anchor='c')
    trv.column("1", width=50, anchor='c')
    trv.column("2", width=50, anchor='c')
    trv.column("3", width=40, anchor='c')
    trv.heading("#0", text='المنتجات', anchor='c')
    trv.heading("1", text='السعر', anchor='c')
    trv.heading("2", text='العدد', anchor='c')
    trv.heading("3", text='حساب الكلي', anchor='c')

    root.mainloop()


title = Label(pro, text="PC Market System", fg="gold",
              bg="black", font=('tajawal', 16, 'bold'))

title.pack(fill=X)

u1 = 'https://www.facebook.com/obourinstitutes'

name1 = ('mohammed', '01279728339', '20 El_Ensaary ST')


def open1():
    webbrowser.open_new(u1)


def about1():
    messagebox.showinfo(
        'developed by', 'Mohammed Mahmoud\nYassen Barakat \nKaream Elnagaar \nYossef Mahmoud \Eslam Ahmed')


def about2():
    messagebox.showinfo(
        'About Program', 'this project for selling all pc products using tkinter library')


def log():
    user = En1.get()
    passw = En2.get()
    if user == name1[0] and passw == '12345':
        open()

    else:
        messagebox.showerror(
            'Eror', 'incorrect information\nplease try again :( ')


F1 = Frame(pro, width=230, height=420, bg='#0B2F3A',)
F1.place(x=570, y=37)
Title1 = Label(F1, text='متجر الكترونيات',
               bg='#0B2F3A', fg='white', font=('tajawal', 12, 'bold'))
Title1.place(x=65, y=10)
Title2 = Label(F1, text='علوم حاسب فرقه ثانيه', bg='#0B2F3A',
               fg='white', font=('tajawal', 12, 'bold'))
Title2.place(x=35, y=50)
Title3 = Label(F1, text='وسائل للتواصل بنا', bg='#0B2F3A',
               fg='white', font=('tajawal', 12, 'bold'))
Title3.place(x=52, y=90)

B1 = Button(F1, text='حسابنا علي الفيسبوك',
            width=26, fg='black', bg='#DBA901', font=('tajawal', 11, 'bold'), command=open1)
B1.place(x=7, y=130)
B2 = Button(F1, text='لمحه عن المشروع',
            width=26, fg='black', bg='#DBA901', font=('tajawal', 11, 'bold'), command=about2)
B2.place(x=7, y=177)
B3 = Button(F1, text='لمحه عن المطورين',
            width=26, fg='black', bg='#DBA901', font=('tajawal', 11, 'bold'), command=about1)
B3.place(x=7, y=225)
B6 = Button(F1, text='اغلاق البرنامج',
            width=26, fg='black', bg='red', font=('tajawal', 11, 'bold'), command=quit)
B6.place(x=7, y=356)

photo = PhotoImage(file='D:\orange.png')
imo = Label(pro, image=photo)
imo.place(x=120, y=43, width=308, height=272)
F2 = Frame(pro, width=570, height=120, bg='#0B2F3A')
F2.place(x=0, y=330)
photo1 = PhotoImage(file='D:\logen.png')
imo1 = Label(pro, image=photo1)
imo1.place(x=460, y=340, width=110, height=100)
L1 = Label(F2, text='اسم المستخدم', fg='gold',
           bg='#0B2F3A', font=('tajawal', 16))
L1.place(x=320, y=24)
L2 = Label(F2, text='كلمة المرور', fg='gold',
           bg='#0B2F3A', font=('tajawal', 16))
L2.place(x=355, y=70)
En1 = Entry(F2, font=('tajawal', 13), justify='center')
En1.place(x=135, y=26)
En2 = Entry(F2, font=('tajawal', 13), justify='center', show='*')
En2.place(x=135, y=71)
BL = Button(F2, text='تسجيل الدخول', bg='#DBA901',
            font=('tajawal', 12), width=12, height=3, command=log)
BL.place(x=10, y=23)


pro.mainloop()
